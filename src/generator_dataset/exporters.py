from __future__ import annotations

import json
from io import StringIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pandas.api.types import (
    is_bool_dtype,
    is_datetime64_any_dtype,
    is_float_dtype,
    is_integer_dtype,
    is_numeric_dtype,
)

try:
    import sqlite3
except ModuleNotFoundError:  # pragma: no cover - Linux CI fallback when stdlib sqlite is unavailable
    from pysqlite3 import dbapi2 as sqlite3

from generator_dataset.settings import GenerationContext
from generator_dataset.reports import ReportDefinition, load_report_catalog


ANOMALY_LOG_COLUMNS = [
    "anomaly_type",
    "table_name",
    "primary_key_value",
    "fiscal_year",
    "description",
    "expected_detection_test",
]

DATE_FORMAT = "yyyy-mm-dd"
TIME_FORMAT = "hh:mm"
DATETIME_FORMAT = "yyyy-mm-dd hh:mm"
DECIMAL_FORMAT = "0.00"
INTEGER_FORMAT = "0"
PERCENT_FORMAT = "0.00%"
EXCEL_TABLE_STYLE = "TableStyleMedium2"
TIME_ONLY_COLUMNS = {"StartTime", "EndTime"}
DATETIME_COLUMNS = {"ClockInTime", "ClockOutTime"}


def anomaly_log_dataframe(context: GenerationContext) -> pd.DataFrame:
    if context.anomaly_log:
        return pd.DataFrame(context.anomaly_log, columns=ANOMALY_LOG_COLUMNS)
    return pd.DataFrame(columns=ANOMALY_LOG_COLUMNS)


def _stringify_value(value: object) -> str:
    if isinstance(value, (dict, list)):
        return json.dumps(value, default=str, ensure_ascii=True)
    if value is None:
        return ""
    return str(value)


def _final_validation_stage(context: GenerationContext) -> tuple[str | None, dict[str, object] | None]:
    if not context.validation_results:
        return None, None
    stage_name = next(reversed(context.validation_results))
    details = context.validation_results.get(stage_name)
    return stage_name, details if isinstance(details, dict) else None


def build_overview_dataframe(context: GenerationContext) -> pd.DataFrame:
    final_stage, final_results = _final_validation_stage(context)
    final_exception_count = len(final_results.get("exceptions", [])) if final_results else 0
    total_rows = sum(len(df) for df in context.tables.values())

    rows = [
        {"Metric": "CompanyName", "Value": context.settings.company_name},
        {"Metric": "FiscalYearStart", "Value": context.settings.fiscal_year_start},
        {"Metric": "FiscalYearEnd", "Value": context.settings.fiscal_year_end},
        {"Metric": "RandomSeed", "Value": context.settings.random_seed},
        {"Metric": "AnomalyMode", "Value": context.settings.anomaly_mode},
        {"Metric": "DatasetTableCount", "Value": len(context.tables)},
        {"Metric": "DatasetRowCount", "Value": total_rows},
        {"Metric": "AnomalyCount", "Value": len(context.anomaly_log)},
        {"Metric": "FinalValidationStage", "Value": final_stage or ""},
        {"Metric": "FinalValidationExceptionCount", "Value": final_exception_count},
        {"Metric": "SQLitePath", "Value": context.settings.sqlite_path if context.settings.export_sqlite else ""},
        {"Metric": "ExcelPath", "Value": context.settings.excel_path if context.settings.export_excel else ""},
        {
            "Metric": "SupportWorkbookPath",
            "Value": context.settings.support_excel_path if context.settings.export_support_excel else "",
        },
        {"Metric": "CsvZipPath", "Value": context.settings.csv_zip_path if context.settings.export_csv_zip else ""},
        {"Metric": "GenerationLogPath", "Value": context.settings.generation_log_path},
    ]

    return pd.DataFrame(rows, columns=["Metric", "Value"])


def validation_stages_dataframe(context: GenerationContext) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for stage, details in context.validation_results.items():
        if not isinstance(details, dict):
            continue
        check_count = 0
        checks_with_exceptions = 0
        nested_exception_total = 0
        for value in details.values():
            if isinstance(value, dict) and "exception_count" in value:
                check_count += 1
                exception_count = int(value.get("exception_count", 0) or 0)
                nested_exception_total += exception_count
                if exception_count > 0:
                    checks_with_exceptions += 1

        rows.append({
            "Stage": stage,
            "ValidationScope": details.get("validation_scope", ""),
            "DirectExceptionCount": len(details.get("exceptions", [])),
            "NestedCheckCount": check_count,
            "ChecksWithExceptions": checks_with_exceptions,
            "NestedExceptionCount": nested_exception_total,
            "AnomalyCount": details.get("anomaly_count", ""),
        })

    return pd.DataFrame(rows)


def validation_checks_dataframe(context: GenerationContext) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for stage, details in context.validation_results.items():
        if not isinstance(details, dict):
            continue
        for area, value in details.items():
            if not isinstance(value, dict) or area == "row_counts":
                continue
            scalar_metrics = {
                metric_key: metric_value
                for metric_key, metric_value in value.items()
                if not isinstance(metric_value, (dict, list))
            }
            rows.append({
                "Stage": stage,
                "Area": area,
                "ExceptionCount": value.get("exception_count", ""),
                "ScalarMetrics": json.dumps(scalar_metrics, default=str, ensure_ascii=True) if scalar_metrics else "",
            })

    return pd.DataFrame(rows)


def _flatten_exception_row(stage: str, area: str, exception_index: int, exception: object) -> dict[str, object]:
    if isinstance(exception, dict):
        exception_type = (
            exception.get("type")
            or exception.get("anomaly_type")
            or exception.get("message")
            or exception.get("description")
            or "dict_exception"
        )
        message = exception.get("message") or exception.get("description") or str(exception_type)
        detail = json.dumps(exception, default=str, ensure_ascii=True)
    else:
        exception_type = type(exception).__name__
        message = str(exception)
        detail = _stringify_value(exception)

    return {
        "Stage": stage,
        "Area": area,
        "ExceptionIndex": exception_index,
        "ExceptionType": exception_type,
        "Message": message,
        "Detail": detail,
    }


def validation_exceptions_dataframe(context: GenerationContext) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for stage, details in context.validation_results.items():
        if not isinstance(details, dict):
            continue

        for index, exception in enumerate(details.get("exceptions", []), start=1):
            rows.append(_flatten_exception_row(stage, "direct", index, exception))

        for area, value in details.items():
            if not isinstance(value, dict):
                continue
            for index, exception in enumerate(value.get("exceptions", []), start=1):
                rows.append(_flatten_exception_row(stage, area, index, exception))

    return pd.DataFrame(rows)


def _looks_like_time_only(series: pd.Series) -> bool:
    non_null = series.dropna()
    if non_null.empty:
        return False
    as_text = non_null.astype(str).str.strip()
    return as_text.str.fullmatch(r"\d{1,2}:\d{2}(:\d{2})?").all()


def _parse_time_series(series: pd.Series) -> pd.Series | None:
    non_null_count = int(series.notna().sum())
    if non_null_count == 0 or not _looks_like_time_only(series):
        return None

    parsed = pd.to_datetime(series, format="%H:%M:%S", errors="coerce")
    if int(parsed.notna().sum()) != non_null_count:
        parsed = pd.to_datetime(series, format="%H:%M", errors="coerce")
    if int(parsed.notna().sum()) != non_null_count:
        return None
    return parsed.dt.time


def _parse_datetime_series(series: pd.Series) -> pd.Series | None:
    non_null_count = int(series.notna().sum())
    if non_null_count == 0:
        return None
    parsed = pd.to_datetime(series, errors="coerce")
    if int(parsed.notna().sum()) != non_null_count:
        return None
    return parsed


def _prepared_excel_dataframe(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, str]]:
    prepared = df.copy()
    formats: dict[str, str] = {}

    for column in prepared.columns:
        series = prepared[column]
        lower_name = str(column).lower()

        if column in TIME_ONLY_COLUMNS:
            parsed_time = _parse_time_series(series)
            if parsed_time is not None:
                prepared[column] = parsed_time
                formats[str(column)] = TIME_FORMAT
                continue

        if column in DATETIME_COLUMNS:
            parsed_datetime = _parse_datetime_series(series)
            if parsed_datetime is not None:
                prepared[column] = parsed_datetime
                formats[str(column)] = DATETIME_FORMAT
                continue

        if "date" in lower_name:
            parsed_date = _parse_datetime_series(series)
            if parsed_date is not None:
                prepared[column] = parsed_date
                formats[str(column)] = DATE_FORMAT
                continue

        if lower_name.endswith("time"):
            parsed_time = _parse_time_series(series)
            if parsed_time is not None:
                prepared[column] = parsed_time
                formats[str(column)] = TIME_FORMAT
                continue

            parsed_datetime = _parse_datetime_series(series)
            if parsed_datetime is not None:
                prepared[column] = parsed_datetime
                formats[str(column)] = DATETIME_FORMAT
                continue

        if is_datetime64_any_dtype(prepared[column]):
            formats[str(column)] = DATETIME_FORMAT
            continue

        if is_bool_dtype(series):
            continue

        if "pct" in lower_name or "percent" in lower_name:
            if is_numeric_dtype(series):
                formats[str(column)] = PERCENT_FORMAT
            continue

        if is_integer_dtype(series):
            formats[str(column)] = INTEGER_FORMAT
            continue

        if is_float_dtype(series) or is_numeric_dtype(series):
            formats[str(column)] = DECIMAL_FORMAT

    return prepared, formats


def _auto_column_width(worksheet, dataframe: pd.DataFrame) -> None:
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        sample = dataframe[column_name].head(200)
        sample_lengths = [len(str(column_name))]
        sample_lengths.extend(len(str(value)) for value in sample if value is not None)
        width = min(max(sample_lengths) + 2, 40)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _table_name(base_name: str, index: int) -> str:
    sanitized = "".join(character if character.isalnum() else "_" for character in base_name)
    return f"T{index}_{sanitized}"[:255]


def _format_worksheet_as_table(worksheet, dataframe: pd.DataFrame, formats: dict[str, str], table_name: str) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = True

    if worksheet.max_column == 0:
        return

    for header_cell in worksheet[1]:
        header_cell.value = "" if header_cell.value is None else str(header_cell.value)

    last_column = get_column_letter(worksheet.max_column)
    table = Table(displayName=table_name, ref=f"A1:{last_column}{worksheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name=EXCEL_TABLE_STYLE,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    for column_index, header_cell in enumerate(worksheet[1], start=1):
        header_cell.alignment = Alignment(horizontal="center")
        number_format = formats.get(str(header_cell.value))
        if not number_format or worksheet.max_row < 2:
            continue
        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = number_format

    _auto_column_width(worksheet, dataframe)


def _write_excel_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for index, (sheet_name, df) in enumerate(sheets.items(), start=1):
            prepared, formats = _prepared_excel_dataframe(df)
            prepared.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            worksheet = writer.sheets[sheet_name[:31]]
            _format_worksheet_as_table(worksheet, prepared, formats, _table_name(sheet_name, index))


def _support_workbook_sheets(context: GenerationContext) -> dict[str, pd.DataFrame]:
    return {
        "Overview": build_overview_dataframe(context),
        "AnomalyLog": anomaly_log_dataframe(context),
        "ValidationStages": validation_stages_dataframe(context),
        "ValidationChecks": validation_checks_dataframe(context),
        "ValidationExceptions": validation_exceptions_dataframe(context),
    }


def export_sqlite(context: GenerationContext) -> None:
    path = Path(context.settings.sqlite_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(path) as connection:
        for table_name, df in context.tables.items():
            df.to_sql(table_name, connection, if_exists="replace", index=False)


def export_excel(context: GenerationContext) -> None:
    dataset_sheets = {table_name[:31]: df for table_name, df in context.tables.items()}
    _write_excel_workbook(Path(context.settings.excel_path), dataset_sheets)


def export_support_excel(context: GenerationContext) -> None:
    _write_excel_workbook(Path(context.settings.support_excel_path), _support_workbook_sheets(context))


def export_csv_zip(context: GenerationContext) -> None:
    path = Path(context.settings.csv_zip_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    with ZipFile(path, mode="w", compression=ZIP_DEFLATED) as archive:
        for table_name, df in context.tables.items():
            buffer = StringIO()
            df.to_csv(buffer, index=False)
            archive.writestr(f"{table_name}.csv", buffer.getvalue().encode("utf-8"))


def _json_safe_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except TypeError:
            pass
    if isinstance(value, (dict, list)):
        return json.loads(json.dumps(value, default=str, ensure_ascii=True))
    if hasattr(value, "item"):
        try:
            return value.item()
        except ValueError:
            pass
    return value


def _report_asset_dir(context: GenerationContext, report: ReportDefinition) -> Path:
    return Path(context.settings.report_output_dir).joinpath(*report.asset_parts)


def _report_preview_payload(
    context: GenerationContext,
    report: ReportDefinition,
    frame: pd.DataFrame,
    *,
    generated_at: str,
) -> dict[str, object]:
    limit = min(report.preview_row_limit, context.settings.report_preview_row_count)
    preview_frame = frame.head(limit)

    rows = []
    for row in preview_frame.to_dict(orient="records"):
        rows.append({key: _json_safe_value(value) for key, value in row.items()})

    return {
        "slug": report.slug,
        "title": report.title,
        "area": report.area,
        "processGroup": report.process_group,
        "cadence": report.cadence,
        "description": report.description,
        "generatedAt": generated_at,
        "rowCount": int(len(frame.index)),
        "previewRowCount": int(len(preview_frame.index)),
        "previewRowLimit": limit,
        "columns": [str(column) for column in frame.columns],
        "rows": rows,
    }


def _write_report_excel(path: Path, report: ReportDefinition, frame: pd.DataFrame) -> None:
    sheet_name = report.title[:31] or "Report"
    _write_excel_workbook(path, {sheet_name: frame})


def export_reports(context: GenerationContext) -> None:
    catalog = load_report_catalog()
    generated_at = pd.Timestamp.now(tz="UTC").isoformat()
    sqlite_path = Path(context.settings.sqlite_path)

    report_root = Path(context.settings.report_output_dir)
    report_root.mkdir(parents=True, exist_ok=True)

    with sqlite3.connect(sqlite_path) as connection:
        for report in catalog:
            frame = pd.read_sql_query(report.query_file.read_text(encoding="utf-8"), connection)
            asset_dir = _report_asset_dir(context, report)
            asset_dir.mkdir(parents=True, exist_ok=True)

            if report.excel_enabled:
                _write_report_excel(asset_dir / f"{report.slug}.xlsx", report, frame)

            if report.csv_enabled:
                frame.to_csv(asset_dir / f"{report.slug}.csv", index=False)

            preview_payload = _report_preview_payload(
                context,
                report,
                frame,
                generated_at=generated_at,
            )
            (asset_dir / "preview.json").write_text(
                json.dumps(preview_payload, indent=2, ensure_ascii=True),
                encoding="utf-8",
            )


def export_validation_report(context: GenerationContext) -> None:
    # Deprecated compatibility shim: standalone validation JSON is no longer emitted.
    _ = context
