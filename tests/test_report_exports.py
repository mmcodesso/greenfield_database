from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook

from generator_dataset.reports import load_report_catalog
from generator_dataset.settings import load_settings


EXPECTED_REPORT_SLUGS = [
    "monthly-income-statement",
    "monthly-balance-sheet",
    "monthly-indirect-cash-flow",
    "pro-forma-monthly-income-statement",
    "pro-forma-monthly-balance-sheet",
    "pro-forma-monthly-indirect-cash-flow",
    "budget-vs-actual-statement-bridge-monthly",
    "monthly-revenue-and-gross-margin",
    "ar-aging",
    "ap-aging",
    "monthly-ar-aging-detail",
    "monthly-ar-aging-summary",
    "monthly-ap-aging-detail",
    "monthly-ap-aging-summary",
    "working-capital-bridge-by-month",
    "budget-vs-actual-working-capital-and-cash-bridge",
    "customer-credit-and-refunds",
    "customer-deposits-and-unapplied-cash-aging",
    "cash-conversion-timing-review",
    "price-realization-vs-list-by-customer-and-portfolio",
    "payroll-expense-mix-by-cost-center-and-pay-class",
    "payroll-and-people-cost-mix-by-cost-center-job-family-level",
    "gross-to-net-payroll-review",
    "hourly-payroll-hours-to-paid-earnings-bridge",
    "payroll-cash-payments-and-remittances",
    "payroll-liability-rollforward",
    "budget-vs-actual-by-cost-center",
    "budget-vs-actual-revenue-price-volume-cost-bridge",
    "sales-and-margin-by-collection-and-style",
    "monthly-work-center-utilization",
    "headcount-by-cost-center-and-job-family",
    "customer-sales-mix-by-region-and-item-group",
    "supplier-purchasing-activity-by-category",
    "supplier-lead-time-and-receipt-reliability",
    "inventory-coverage-and-projected-stockout-risk",
    "rough-cut-capacity-load-vs-available-hours",
    "forecast-error-and-bias-by-collection-and-style-family",
    "labor-and-headcount-by-work-location-job-family-cost-center",
    "absence-rate-by-work-location-job-family-month",
    "overtime-approval-coverage-and-concentration",
    "punch-to-pay-bridge-for-hourly-workers",
    "approval-and-sod-review",
    "potential-anomaly-review",
    "payroll-control-review",
    "paid-without-clock-and-clock-without-pay-review",
]

EXPECTED_REPORT_COLUMNS = {
    "pro-forma-monthly-income-statement": [
        "FiscalYear",
        "FiscalPeriod",
        "StatementSection",
        "LineLabel",
        "LineType",
        "DisplayOrder",
        "Amount",
    ],
    "pro-forma-monthly-balance-sheet": [
        "FiscalYear",
        "FiscalPeriod",
        "StatementSection",
        "LineLabel",
        "LineType",
        "DisplayOrder",
        "Amount",
    ],
    "pro-forma-monthly-indirect-cash-flow": [
        "FiscalYear",
        "FiscalPeriod",
        "StatementSection",
        "LineLabel",
        "LineType",
        "DisplayOrder",
        "Amount",
    ],
    "budget-vs-actual-statement-bridge-monthly": [
        "FiscalYear",
        "FiscalPeriod",
        "StatementName",
        "LineLabel",
        "BudgetAmount",
        "ActualAmount",
        "VarianceAmount",
        "VariancePct",
    ],
    "monthly-ar-aging-detail": [
        "MonthEndDate",
        "InvoiceMonth",
        "InvoiceNumber",
        "CustomerName",
        "Region",
        "CustomerSegment",
        "InvoiceDate",
        "DueDate",
        "DaysFromDueAtMonthEnd",
        "AgingBucket",
        "InvoiceAmount",
        "CashAppliedAsOfMonthEnd",
        "CreditMemoAppliedAsOfMonthEnd",
        "OpenAmountAsOfMonthEnd",
    ],
    "monthly-ar-aging-summary": [
        "MonthEndDate",
        "CustomerName",
        "Region",
        "CustomerSegment",
        "OpenInvoiceCount",
        "TotalOpenAmount",
        "CurrentAmount",
        "Days1To30Amount",
        "Days31To60Amount",
        "Days61To90Amount",
        "Days90PlusAmount",
        "PastDueAmount",
        "OldestDaysPastDue",
    ],
    "monthly-ap-aging-detail": [
        "MonthEndDate",
        "InvoiceMonth",
        "InvoiceNumber",
        "SupplierName",
        "SupplierCategory",
        "SupplierRiskRating",
        "InvoiceDate",
        "DueDate",
        "DaysFromDueAtMonthEnd",
        "AgingBucket",
        "InvoiceAmount",
        "CashPaidAsOfMonthEnd",
        "OpenAmountAsOfMonthEnd",
    ],
    "monthly-ap-aging-summary": [
        "MonthEndDate",
        "SupplierName",
        "SupplierCategory",
        "SupplierRiskRating",
        "OpenInvoiceCount",
        "TotalOpenAmount",
        "CurrentAmount",
        "Days1To30Amount",
        "Days31To60Amount",
        "Days61To90Amount",
        "Days90PlusAmount",
        "PastDueAmount",
        "OldestDaysPastDue",
    ],
    "working-capital-bridge-by-month": [
        "FiscalYear",
        "FiscalPeriod",
        "AccountsReceivableEndingBalance",
        "InventoryAndWIPEndingBalance",
        "AccountsPayableEndingBalance",
        "GRNIEndingBalance",
        "CustomerDepositsEndingBalance",
        "AccruedExpensesEndingBalance",
        "PayrollLiabilitiesEndingBalance",
        "NetWorkingCapital",
    ],
    "budget-vs-actual-working-capital-and-cash-bridge": [
        "FiscalYear",
        "FiscalPeriod",
        "MetricName",
        "BudgetAmount",
        "ActualAmount",
        "VarianceAmount",
        "VariancePct",
    ],
    "customer-credit-and-refunds": [
        "CreditMemoNumber",
        "CreditMemoDate",
        "CustomerName",
        "OriginalInvoiceNumber",
        "CreditMemoAmount",
        "RemainingARBeforeMemo",
        "CustomerCreditCreated",
        "RefundedAmount",
        "OpenCustomerCredit",
    ],
    "customer-deposits-and-unapplied-cash-aging": [
        "ReceiptNumber",
        "ReceiptDate",
        "CustomerName",
        "ReceiptAmount",
        "AppliedAmount",
        "OpenUnappliedAmount",
        "AppliedInvoiceCount",
        "FirstApplicationDate",
        "LastApplicationDate",
        "AgeDaysAtSnapshot",
        "DaysToFirstApplication",
    ],
    "cash-conversion-timing-review": [
        "MetricFamily",
        "FiscalYear",
        "FiscalPeriod",
        "DocumentCount",
        "SettledDocumentCount",
        "OpenDocumentCount",
        "AvgDaysToFirstSettlement",
        "MaxDaysToFirstSettlement",
    ],
    "price-realization-vs-list-by-customer-and-portfolio": [
        "InvoiceMonth",
        "Region",
        "CustomerSegment",
        "CustomerName",
        "CollectionName",
        "StyleFamily",
        "InvoicedQuantity",
        "BaseListRevenue",
        "NetRevenue",
        "AveragePromotionDiscountPct",
        "PriceRealizationPct",
    ],
    "payroll-expense-mix-by-cost-center-and-pay-class": [
        "PayMonth",
        "CostCenterName",
        "PayClass",
        "EmployeeCount",
        "GrossPay",
        "EmployeeWithholdings",
        "EmployerPayrollTax",
        "EmployerBenefits",
        "NetPay",
    ],
    "payroll-and-people-cost-mix-by-cost-center-job-family-level": [
        "CostCenterName",
        "JobFamily",
        "JobLevel",
        "PayClass",
        "EndStateHeadcount",
        "ActiveHeadcount",
        "TerminatedHeadcount",
        "EmployeesWithPayroll",
        "GrossPay",
        "EmployerPayrollTax",
        "EmployerBenefits",
        "TotalPeopleCost",
    ],
    "gross-to-net-payroll-review": [
        "PeriodNumber",
        "PayDate",
        "EmployeeID",
        "EmployeeName",
        "CostCenterName",
        "PayClass",
        "GrossPay",
        "EmployeeWithholdings",
        "EmployerPayrollTax",
        "EmployerBenefits",
        "NetPay",
        "GrossLessWithholdings",
        "EmployerBurden",
        "Status",
    ],
    "hourly-payroll-hours-to-paid-earnings-bridge": [
        "PeriodNumber",
        "PayDate",
        "EmployeeID",
        "EmployeeName",
        "CostCenterName",
        "ApprovedRegularHours",
        "ApprovedOvertimeHours",
        "LaborHoursSupported",
        "PayrollRegularHours",
        "PayrollOvertimeHours",
        "HourlyEarningsAmount",
        "GrossPay",
        "NetPay",
        "FirstPaymentDate",
        "Status",
    ],
    "payroll-cash-payments-and-remittances": [
        "FiscalYear",
        "FiscalPeriod",
        "NetPayCash",
        "EmployeeTaxRemittance",
        "EmployerTaxRemittance",
        "BenefitsRemittance",
        "TotalLiabilityRemittances",
        "TotalPayrollCashOutflow",
    ],
    "payroll-liability-rollforward": [
        "FiscalYear",
        "FiscalPeriod",
        "AccountNumber",
        "AccountName",
        "DebitAmount",
        "CreditAmount",
        "NetIncrease",
        "EndingBalance",
    ],
    "customer-sales-mix-by-region-and-item-group": [
        "InvoiceMonth",
        "Region",
        "CustomerSegment",
        "ItemGroup",
        "ItemCode",
        "ItemName",
        "BilledQuantity",
        "RevenueAmount",
        "AverageUnitPrice",
    ],
    "supplier-purchasing-activity-by-category": [
        "OrderMonth",
        "SupplierCategory",
        "SupplierRiskRating",
        "SupplierName",
        "ItemGroup",
        "PurchaseOrderCount",
        "PurchaseOrderLineCount",
        "OrderedQuantity",
        "OrderedValue",
    ],
    "supplier-lead-time-and-receipt-reliability": [
        "OrderMonth",
        "SupplierName",
        "SupplierCategory",
        "PurchaseOrderCount",
        "AvgDaysToFirstReceipt",
        "AvgDaysToFullReceipt",
        "FullyReceivedPOCount",
        "PartiallyReceivedPOCount",
        "NoReceiptPOCount",
        "FirstReceiptOver14DaysCount",
    ],
    "inventory-coverage-and-projected-stockout-risk": [
        "ItemCode",
        "ItemName",
        "ItemGroup",
        "CollectionName",
        "StyleFamily",
        "WarehouseName",
        "AvgWeeklyForecastQuantity",
        "ProjectedAvailableQuantity",
        "NetRequirementQuantity",
        "RecommendedOrderQuantity",
        "WeeksOfCoverage",
        "StockoutRisk",
    ],
    "rough-cut-capacity-load-vs-available-hours": [
        "BucketWeekStartDate",
        "BucketWeekEndDate",
        "WorkCenterCode",
        "WorkCenterName",
        "PlannedLoadHours",
        "AvailableHours",
        "UtilizationPct",
        "CapacityStatus",
    ],
    "forecast-error-and-bias-by-collection-and-style-family": [
        "CollectionName",
        "StyleFamily",
        "ForecastQuantity",
        "ActualOrderQuantity",
        "BiasQuantity",
        "AbsoluteErrorQuantity",
    ],
    "labor-and-headcount-by-work-location-job-family-cost-center": [
        "WorkLocation",
        "CostCenterName",
        "JobFamily",
        "EndStateHeadcount",
        "ActiveHeadcount",
        "TerminatedHeadcount",
        "GrossPay",
        "EmployerBurden",
        "TotalPeopleCost",
        "ApprovedClockHours",
        "DirectLaborHours",
        "ExtendedLaborCost",
    ],
    "absence-rate-by-work-location-job-family-month": [
        "YearMonth",
        "WorkLocation",
        "JobFamily",
        "RosteredHours",
        "HoursAbsent",
        "PaidAbsenceHours",
        "UnpaidAbsenceHours",
        "AbsenceRate",
    ],
    "overtime-approval-coverage-and-concentration": [
        "YearMonth",
        "WorkCenterCode",
        "WorkCenterName",
        "OvertimeHours",
        "ApprovedOvertimeHours",
        "OvertimeEntryCount",
        "ApprovedOvertimeEntryCount",
        "MissingApprovalEntryCount",
        "OvertimeApprovalCoveragePct",
    ],
    "punch-to-pay-bridge-for-hourly-workers": [
        "PeriodNumber",
        "PeriodStartDate",
        "PeriodEndDate",
        "EmployeeNumber",
        "EmployeeName",
        "JobTitle",
        "JobFamily",
        "PunchCount",
        "ApprovedClockHours",
        "ApprovedOvertimeHours",
        "LaborHours",
        "LaborCost",
        "GrossPay",
        "NetPay",
    ],
    "budget-vs-actual-revenue-price-volume-cost-bridge": [
        "FiscalYear",
        "FiscalPeriod",
        "CollectionName",
        "StyleFamily",
        "BudgetQuantity",
        "ActualQuantity",
        "BudgetNetRevenue",
        "ActualNetRevenue",
        "BudgetCOGS",
        "ActualCOGS",
        "RevenueVariance",
        "PriceVariance",
        "VolumeVariance",
        "CostVariance",
    ],
    "payroll-control-review": [
        "PotentialIssue",
        "ReferenceNumber",
        "EmployeeReference",
        "EmployeeName",
        "EventDate",
        "Amount",
    ],
    "paid-without-clock-and-clock-without-pay-review": [
        "PotentialIssue",
        "ReferenceNumber",
        "EmployeeName",
        "EventDate",
        "HoursOrAmount",
    ],
}


def _header_row(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    try:
        return [str(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    finally:
        workbook.close()


def test_report_catalog_contains_curated_report_library() -> None:
    catalog = load_report_catalog()
    assert [report.slug for report in catalog] == EXPECTED_REPORT_SLUGS


def test_report_exports_create_curated_artifacts(
    report_validation_dataset_artifacts: dict[str, object],
) -> None:
    report_root = Path(report_validation_dataset_artifacts["report_output_dir"])

    for report in load_report_catalog():
        asset_dir = report_root / report.area / report.process_group / report.slug
        excel_path = asset_dir / f"{report.slug}.xlsx"
        csv_path = asset_dir / f"{report.slug}.csv"
        preview_path = asset_dir / "preview.json"

        assert asset_dir.exists(), f"Missing asset directory for {report.slug}"
        assert excel_path.exists() and excel_path.stat().st_size > 0
        assert csv_path.exists() and csv_path.stat().st_size > 0
        assert preview_path.exists() and preview_path.stat().st_size > 0


def test_report_exports_preserve_columns_and_preview_schema(
    report_validation_dataset_artifacts: dict[str, object],
) -> None:
    context = report_validation_dataset_artifacts["context"]
    report_root = Path(report_validation_dataset_artifacts["report_output_dir"])

    for report in load_report_catalog():
        asset_dir = report_root / report.area / report.process_group / report.slug
        csv_frame = pd.read_csv(asset_dir / f"{report.slug}.csv")
        preview = json.loads((asset_dir / "preview.json").read_text(encoding="utf-8"))
        workbook_headers = _header_row(asset_dir / f"{report.slug}.xlsx")
        effective_preview_limit = min(report.preview_row_limit, context.settings.report_preview_row_count)

        assert not csv_frame.empty, f"Expected rows in exported report {report.slug}"
        assert preview["slug"] == report.slug
        assert preview["title"] == report.title
        assert preview["area"] == report.area
        assert preview["processGroup"] == report.process_group
        assert preview["rowCount"] == len(csv_frame.index)
        assert preview["previewRowLimit"] == effective_preview_limit
        assert preview["previewRowCount"] == min(len(csv_frame.index), effective_preview_limit)
        assert preview["columns"] == csv_frame.columns.tolist()
        assert preview["columns"] == workbook_headers
        assert len(preview["rows"]) == preview["previewRowCount"]
        assert preview["generatedAt"]


def test_report_manifest_and_docs_include_curated_paths() -> None:
    manifest_text = Path("src/generated/reportManifest.js").read_text(encoding="utf-8")
    collections_text = Path("src/generated/reportDocCollections.js").read_text(encoding="utf-8")
    sidebar_text = Path("sidebars.js").read_text(encoding="utf-8")
    analytics_hub = Path("docs/analytics/index.md").read_text(encoding="utf-8")
    reports_hub = Path("docs/analytics/reports/index.md").read_text(encoding="utf-8")
    financial_reports = Path("docs/analytics/reports/financial.md").read_text(encoding="utf-8")
    managerial_reports = Path("docs/analytics/reports/managerial.md").read_text(encoding="utf-8")
    audit_reports = Path("docs/analytics/reports/audit.md").read_text(encoding="utf-8")
    docusaurus_config = Path("docusaurus.config.js").read_text(encoding="utf-8")

    for path in [
        Path("docs/analytics/reports/index.md"),
        Path("docs/analytics/reports/financial.md"),
        Path("docs/analytics/reports/managerial.md"),
        Path("docs/analytics/reports/audit.md"),
    ]:
        assert path.exists(), f"Missing report doc: {path}"

    assert '"analytics/reports/index"' in sidebar_text
    assert '"analytics/reports/financial"' in sidebar_text
    assert '"analytics/reports/managerial"' in sidebar_text
    assert '"analytics/reports/audit"' in sidebar_text
    assert 'staticDirectories: ["static", "queries"]' in docusaurus_config
    assert "Reports Hub" in analytics_hub
    assert "Business Perspectives" in reports_hub
    assert "Report Library" in reports_hub
    assert "reportAreaCollections.financial" in financial_reports
    assert "reportAreaCollections.managerial" in managerial_reports
    assert "reportAreaCollections.audit" in audit_reports
    assert "receivables-and-payables-aging" in collections_text
    assert "customer-and-supplier-analysis" in collections_text
    assert "working-capital-and-cash-conversion" in collections_text
    assert "customer-pricing-and-settlement" in collections_text
    assert "payroll-cost-and-mix" in collections_text
    assert "payroll-cash-and-liabilities" in collections_text
    assert "planning-and-supply-risk" in collections_text
    assert "workforce-and-payroll-operations" in collections_text
    assert "payroll-and-time-controls" in collections_text

    for report in load_report_catalog():
        asset_base_path = f"/reports/{report.area}/{report.process_group}/{report.slug}"
        assert report.slug in manifest_text
        assert report.slug in collections_text
        assert f'"previewPath": "{asset_base_path}/preview.json"' in manifest_text
        assert f'"excelPath": "{asset_base_path}/{report.slug}.xlsx"' in manifest_text
        assert f'"csvPath": "{asset_base_path}/{report.slug}.csv"' in manifest_text


def test_report_settings_require_sqlite_export(tmp_path: Path) -> None:
    config_path = tmp_path / "settings.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "random_seed": 20260401,
                "fiscal_year_start": "2026-01-01",
                "fiscal_year_end": "2026-12-31",
                "company_name": "Charles River Home Furnishings, Inc.",
                "short_name": "CharlesRiver",
                "base_url": "https://charlesriver.accountinganalyticshub.com",
                "tax_rate": 0.065,
                "employee_count": 48,
                "customer_count": 60,
                "supplier_count": 35,
                "item_count": 90,
                "warehouse_count": 2,
                "export_sqlite": False,
                "export_excel": False,
                "export_support_excel": False,
                "export_csv_zip": False,
                "export_reports": True,
                "anomaly_mode": "none",
                "sqlite_path": "outputs/{short_name}_validation.sqlite",
                "excel_path": "outputs/{short_name}_validation.xlsx",
                "support_excel_path": "outputs/{short_name}_validation_support.xlsx",
                "csv_zip_path": "outputs/{short_name}_validation_csv.zip",
                "report_output_dir": "outputs/reports",
                "report_preview_row_count": 25,
                "generation_log_path": "outputs/generation_validation.log",
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="export_reports requires export_sqlite"):
        load_settings(config_path)


def test_curated_reports_export_expected_columns(
    report_validation_dataset_artifacts: dict[str, object],
) -> None:
    report_root = Path(report_validation_dataset_artifacts["report_output_dir"])

    for report in load_report_catalog():
        expected_columns = EXPECTED_REPORT_COLUMNS.get(report.slug)
        if expected_columns is None:
            continue

        csv_frame = pd.read_csv(
            report_root / report.area / report.process_group / report.slug / f"{report.slug}.csv"
        )
        assert csv_frame.columns.tolist() == expected_columns
