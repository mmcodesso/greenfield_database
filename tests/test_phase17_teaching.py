from __future__ import annotations

from collections import Counter
import sqlite3
from pathlib import Path
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

from generator_dataset.main import build_phase17


TARGETED_AUDIT_QUERY_EXPECTATIONS = {
    "05_duplicate_payment_reference_review.sql": 1,
    "06_potential_anomaly_review.sql": 1,
    "11_payroll_control_review.sql": 1,
    "12_labor_time_after_close_and_paid_without_time.sql": 1,
    "13_over_under_accrual_review.sql": 1,
    "14_missing_routing_or_operation_link_review.sql": 1,
    "15_operation_sequence_and_final_completion_review.sql": 1,
    "16_schedule_on_nonworking_day_review.sql": 1,
    "17_over_capacity_day_review.sql": 1,
    "18_completion_before_scheduled_operation_end.sql": 1,
    "19_time_clock_exceptions_by_employee_supervisor_work_center.sql": 1,
    "20_labor_outside_scheduled_operation_window_review.sql": 1,
    "21_paid_without_clock_and_clock_without_pay_review.sql": 1,
    "25_time_clock_payroll_labor_bridge_review.sql": 1,
    "26_duplicate_ap_reference_detail_review.sql": 1,
}


def _execute_all_starter_sql(sqlite_path: Path) -> None:
    sql_files = sorted(Path("queries").rglob("*.sql"))
    assert sql_files, "No starter SQL files were found."

    with sqlite3.connect(sqlite_path) as connection:
        for sql_file in sql_files:
            sql = sql_file.read_text(encoding="utf-8")
            result = pd.read_sql_query(sql, connection)
            assert len(result.columns) >= 1, f"Query returned no columns: {sql_file}"


def test_phase17_helper_generates_clean_dataset() -> None:
    context = build_phase17("config/settings_validation.yaml", validation_scope="operational")
    phase17 = context.validation_results["phase17"]

    assert phase17["exceptions"] == []
    assert phase17["validation_scope"] == "operational"
    assert phase17["time_clock_controls"]["exception_count"] == 0
    assert phase17["capacity_controls"]["exception_count"] == 0


def test_phase17_new_anomalies_are_logged_and_detectable(
    default_anomaly_dataset_artifacts: dict[str, object],
) -> None:
    context = default_anomaly_dataset_artifacts["context"]
    results = context.validation_results["phase8"]
    anomaly_counts = Counter(entry["anomaly_type"] for entry in context.anomaly_log)

    for anomaly_type in [
        "duplicate_supplier_invoice_number",
        "missing_payroll_payment",
        "payroll_payment_before_approval",
        "missing_work_order_operations",
        "invalid_direct_labor_operation_link",
        "overlapping_operation_sequence",
    ]:
        assert anomaly_counts[anomaly_type] > 0

    duplicate_payment_refs = (
        context.tables["DisbursementPayment"]
        .dropna(subset=["CheckNumber"])
        .groupby(["SupplierID", "CheckNumber"])
        .size()
    )
    assert duplicate_payment_refs.gt(1).any()

    duplicate_invoice_numbers = context.tables["PurchaseInvoice"].groupby(["SupplierID", "InvoiceNumber"]).size()
    assert duplicate_invoice_numbers.gt(1).any()

    payment_register_ids = set(context.tables["PayrollPayment"]["PayrollRegisterID"].astype(int))
    approved_registers = context.tables["PayrollRegister"][
        context.tables["PayrollRegister"]["Status"].eq("Approved")
    ]["PayrollRegisterID"].astype(int)
    assert any(register_id not in payment_register_ids for register_id in approved_registers)

    payroll_payment_join = context.tables["PayrollPayment"].merge(
        context.tables["PayrollRegister"][["PayrollRegisterID", "ApprovedDate"]],
        on="PayrollRegisterID",
        how="inner",
    )
    assert (
        pd.to_datetime(payroll_payment_join["PaymentDate"])
        < pd.to_datetime(payroll_payment_join["ApprovedDate"])
    ).any()

    work_orders = context.tables["WorkOrder"][["WorkOrderID", "WorkOrderNumber"]].copy()
    operation_counts = context.tables["WorkOrderOperation"].groupby("WorkOrderID").size().rename("OperationCount")
    work_orders = work_orders.merge(operation_counts, on="WorkOrderID", how="left").fillna({"OperationCount": 0})
    assert work_orders["OperationCount"].eq(0).any()

    labor_entries = context.tables["LaborTimeEntry"].copy()
    operations = context.tables["WorkOrderOperation"][["WorkOrderOperationID", "WorkOrderID"]].copy()
    labor_with_ops = labor_entries.merge(
        operations,
        on="WorkOrderOperationID",
        how="left",
        suffixes=("", "_operation"),
    )
    direct_labor_gaps = labor_with_ops[
        labor_with_ops["LaborType"].eq("Direct Manufacturing")
        & (
            labor_with_ops["WorkOrderOperationID"].isna()
            | labor_with_ops["WorkOrderID_operation"].isna()
            | labor_with_ops["WorkOrderID"].astype("Int64") != labor_with_ops["WorkOrderID_operation"].astype("Int64")
        )
    ]
    assert not direct_labor_gaps.empty

    ordered_operations = context.tables["WorkOrderOperation"].sort_values(["WorkOrderID", "OperationSequence"])
    prior_end = ordered_operations.groupby("WorkOrderID")["ActualEndDate"].shift(1)
    overlap_rows = ordered_operations[
        prior_end.notna()
        & ordered_operations["ActualStartDate"].notna()
        & (
            pd.to_datetime(ordered_operations["ActualStartDate"])
            < pd.to_datetime(prior_end)
        )
    ]
    assert not overlap_rows.empty

    assert results["anomaly_count"] >= len(context.anomaly_log)


def test_phase17_default_build_targeted_audit_queries_return_rows(
    default_anomaly_dataset_artifacts: dict[str, object],
) -> None:
    sqlite_path = Path(default_anomaly_dataset_artifacts["sqlite_path"])
    assert sqlite_path.exists()

    with sqlite3.connect(sqlite_path) as connection:
        for sql_name, minimum_rows in TARGETED_AUDIT_QUERY_EXPECTATIONS.items():
            sql = Path("queries/audit", sql_name).read_text(encoding="utf-8")
            result = pd.read_sql_query(sql, connection)
            assert len(result) >= minimum_rows, f"Expected rows from {sql_name}, found {len(result)}"


def test_phase17_clean_validation_build_all_starter_sql_executes(
    clean_validation_dataset_artifacts: dict[str, object],
) -> None:
    sqlite_path = Path(clean_validation_dataset_artifacts["sqlite_path"])
    assert sqlite_path.exists()
    _execute_all_starter_sql(sqlite_path)


def test_phase17_default_build_all_starter_sql_executes(
    default_anomaly_dataset_artifacts: dict[str, object],
) -> None:
    sqlite_path = Path(default_anomaly_dataset_artifacts["sqlite_path"])
    assert sqlite_path.exists()
    _execute_all_starter_sql(sqlite_path)


def test_phase17_default_export_artifacts_follow_split_contract(
    default_anomaly_dataset_artifacts: dict[str, object],
) -> None:
    sqlite_path = Path(default_anomaly_dataset_artifacts["sqlite_path"])
    excel_path = Path(default_anomaly_dataset_artifacts["excel_path"])
    support_excel_path = Path(default_anomaly_dataset_artifacts["support_excel_path"])
    csv_zip_path = Path(default_anomaly_dataset_artifacts["csv_zip_path"])
    assert sqlite_path.exists()
    assert excel_path.exists()
    assert support_excel_path.exists()
    assert csv_zip_path.exists()

    with sqlite3.connect(sqlite_path) as connection:
        tables = pd.read_sql_query(
            "SELECT name FROM sqlite_master WHERE type = 'table' AND name IN ('AnomalyLog', 'ValidationSummary')",
            connection,
        )["name"].tolist()
    assert tables == []

    dataset_workbook = load_workbook(excel_path)
    support_workbook = load_workbook(support_excel_path)

    assert "AnomalyLog" not in dataset_workbook.sheetnames
    assert "ValidationStages" not in dataset_workbook.sheetnames
    for sheet_name in ["Account", "SalesOrder", "GLEntry", "TimeClockEntry"]:
        assert len(dataset_workbook[sheet_name].tables) >= 1

    required_support_sheets = {"Overview", "AnomalyLog", "ValidationStages", "ValidationChecks", "ValidationExceptions"}
    assert required_support_sheets.issubset(set(support_workbook.sheetnames))
    for sheet_name in required_support_sheets:
        assert len(support_workbook[sheet_name].tables) >= 1

    with ZipFile(csv_zip_path) as archive:
        zip_members = set(archive.namelist())
    assert zip_members == {f"{table_name}.csv" for table_name in default_anomaly_dataset_artifacts["context"].tables}


def test_phase17_docs_include_cases_matrix_and_subprocess_diagrams() -> None:
    for path in [
        Path("docs/analytics/cases/index.md"),
        Path("docs/analytics/cases/o2c-trace-case.md"),
        Path("docs/analytics/cases/p2p-accrual-settlement-case.md"),
        Path("docs/analytics/cases/manufacturing-labor-cost-case.md"),
        Path("docs/analytics/cases/demand-planning-and-replenishment-case.md"),
        Path("docs/analytics/cases/product-portfolio-and-lifecycle-case.md"),
        Path("docs/analytics/cases/product-portfolio-profitability-case.md"),
        Path("docs/analytics/cases/workforce-coverage-and-attendance-case.md"),
        Path("docs/analytics/cases/working-capital-and-cash-conversion-case.md"),
        Path("docs/analytics/cases/financial-statement-bridge-case.md"),
        Path("docs/analytics/cases/pricing-and-margin-governance-case.md"),
        Path("docs/analytics/cases/audit-exception-lab.md"),
    ]:
        assert path.exists(), f"Missing Phase 17 case doc: {path}"

    audit_doc = Path("docs/analytics/audit.md").read_text(encoding="utf-8")
    assert "## Anomaly Coverage Queries" in audit_doc

    company_story_text = Path("docs/learn-the-business/company-story.md").read_text(encoding="utf-8")
    assert "Charles River" in company_story_text
    assert "greater Boston area" in company_story_text
    assert "buys some finished goods" in company_story_text
    assert "manufactures a selected subset" in company_story_text
    assert "## What Is Still Simplified" not in company_story_text

    process_flows_text = Path("docs/learn-the-business/process-flows.md").read_text(encoding="utf-8")
    assert "| Process | What it means inside the company |" in process_flows_text
    assert "## Process Map" in process_flows_text
    assert "## How Process Becomes Reporting and Analysis" in process_flows_text
    assert "GLEntry" in process_flows_text

    o2c_case_text = Path("docs/analytics/cases/o2c-trace-case.md").read_text(encoding="utf-8")
    assert "## The Problem to Solve" in o2c_case_text
    assert "## What You Need to Develop" in o2c_case_text
    assert "## Step-by-Step Walkthrough" in o2c_case_text
    assert "## Optional Excel Follow-Through" in o2c_case_text
    assert "## Wrap-Up Questions" in o2c_case_text
    assert "cases/01_o2c_line_trace_order_shipment_invoice.sql" in o2c_case_text
    assert "cases/02_o2c_source_to_gl_trace.sql" in o2c_case_text
    assert "## Recommended Query Sequence" not in o2c_case_text

    p2p_case_text = Path("docs/analytics/cases/p2p-accrual-settlement-case.md").read_text(encoding="utf-8")
    assert "## The Problem to Solve" in p2p_case_text
    assert "## What You Need to Develop" in p2p_case_text
    assert "## Step-by-Step Walkthrough" in p2p_case_text
    assert "## Optional Excel Follow-Through" in p2p_case_text
    assert "## Wrap-Up Questions" in p2p_case_text
    assert "cases/03_p2p_invoice_line_trace_receipt_vs_accrual.sql" in p2p_case_text
    assert "cases/04_p2p_accrual_journal_invoice_payment_gl_trace.sql" in p2p_case_text
    assert "financial/12_accrued_expense_rollforward.sql" in p2p_case_text
    assert "financial/13_accrued_vs_invoiced_vs_paid_timing.sql" in p2p_case_text
    assert "audit/23_accrued_service_settlement_exception_review.sql" in p2p_case_text
    assert "## Recommended Query Sequence" not in p2p_case_text

    manufacturing_case_text = Path("docs/analytics/cases/manufacturing-labor-cost-case.md").read_text(encoding="utf-8")
    assert "## The Problem to Solve" in manufacturing_case_text
    assert "## What You Need to Develop" in manufacturing_case_text
    assert "## Step-by-Step Walkthrough" in manufacturing_case_text
    assert "## Optional Excel Follow-Through" in manufacturing_case_text
    assert "## Wrap-Up Questions" in manufacturing_case_text
    assert "cases/05_manufacturing_work_order_operation_trace.sql" in manufacturing_case_text
    assert "cases/06_manufacturing_work_order_close_gl_trace.sql" in manufacturing_case_text
    assert "managerial/24_approved_clock_hours_vs_labor_allocation.sql" in manufacturing_case_text
    assert "managerial/12_direct_labor_by_work_order_and_employee_class.sql" in manufacturing_case_text
    assert "financial/17_manufacturing_cost_component_bridge.sql" in manufacturing_case_text
    assert "audit/20_labor_outside_scheduled_operation_window_review.sql" in manufacturing_case_text
    assert "## Recommended Query Sequence" not in manufacturing_case_text

    demand_planning_case_text = Path("docs/analytics/cases/demand-planning-and-replenishment-case.md").read_text(encoding="utf-8")
    assert "## The Problem to Solve" in demand_planning_case_text
    assert "## What You Need to Develop" in demand_planning_case_text
    assert "## Step-by-Step Walkthrough" in demand_planning_case_text
    assert "## Optional Excel Follow-Through" in demand_planning_case_text
    assert "## Wrap-Up Questions" in demand_planning_case_text
    assert "financial/23_forecast_vs_actual_demand_by_week_item_group_collection_lifecycle.sql" in demand_planning_case_text
    assert "managerial/45_forecast_error_and_bias_by_collection_style_family.sql" in demand_planning_case_text
    assert "managerial/46_supply_plan_driver_mix_by_collection_and_supply_mode.sql" in demand_planning_case_text
    assert "managerial/42_inventory_coverage_and_projected_stockout_risk.sql" in demand_planning_case_text
    assert "managerial/44_expedite_pressure_by_item_family_and_month.sql" in demand_planning_case_text
    assert "financial/24_recommendation_conversion_by_type_priority_planner.sql" in demand_planning_case_text
    assert "managerial/43_rough_cut_capacity_load_vs_available_hours.sql" in demand_planning_case_text
    assert "## Recommended Query Sequence" not in demand_planning_case_text

    product_portfolio_case_text = Path("docs/analytics/cases/product-portfolio-and-lifecycle-case.md").read_text(encoding="utf-8")
    assert "## The Problem to Solve" in product_portfolio_case_text
    assert "## What You Need to Develop" in product_portfolio_case_text
    assert "## Step-by-Step Walkthrough" in product_portfolio_case_text
    assert "## Optional Excel Follow-Through" in product_portfolio_case_text
    assert "## Wrap-Up Questions" in product_portfolio_case_text
    assert "managerial/31_product_portfolio_mix_by_collection_style_lifecycle_supply_mode.sql" in product_portfolio_case_text
    assert "audit/30_item_master_completeness_review.sql" in product_portfolio_case_text
    assert "audit/36_item_status_alignment_review.sql" in product_portfolio_case_text
    assert "managerial/30_sales_margin_by_collection_style_material.sql" in product_portfolio_case_text
    assert "audit/31_discontinued_or_prelaunch_item_activity_review.sql" in product_portfolio_case_text
    assert "## Recommended Query Sequence" not in product_portfolio_case_text

    product_portfolio_profitability_case_text = Path("docs/analytics/cases/product-portfolio-profitability-case.md").read_text(encoding="utf-8")
    assert "## The Problem to Solve" in product_portfolio_profitability_case_text
    assert "## What You Need to Develop" in product_portfolio_profitability_case_text
    assert "## Step-by-Step Walkthrough" in product_portfolio_profitability_case_text
    assert "## Optional Excel Follow-Through" in product_portfolio_profitability_case_text
    assert "## Wrap-Up Questions" in product_portfolio_profitability_case_text
    assert "managerial/31_product_portfolio_mix_by_collection_style_lifecycle_supply_mode.sql" in product_portfolio_profitability_case_text
    assert "financial/21_revenue_and_gross_margin_by_collection_style_lifecycle_supply_mode.sql" in product_portfolio_profitability_case_text
    assert "managerial/32_contribution_margin_by_collection_material_lifecycle_supply_mode.sql" in product_portfolio_profitability_case_text
    assert "managerial/33_customer_service_impact_by_collection_style.sql" in product_portfolio_profitability_case_text
    assert "managerial/35_portfolio_return_refund_impact_by_collection_lifecycle.sql" in product_portfolio_profitability_case_text
    assert "## Recommended Query Sequence" not in product_portfolio_profitability_case_text

    workforce_coverage_case_text = Path("docs/analytics/cases/workforce-coverage-and-attendance-case.md").read_text(encoding="utf-8")
    assert "## The Problem to Solve" in workforce_coverage_case_text
    assert "## What You Need to Develop" in workforce_coverage_case_text
    assert "## Step-by-Step Walkthrough" in workforce_coverage_case_text
    assert "## Optional Excel Follow-Through" in workforce_coverage_case_text
    assert "## Wrap-Up Questions" in workforce_coverage_case_text
    assert "managerial/36_staffing_coverage_vs_work_center_planned_load.sql" in workforce_coverage_case_text
    assert "managerial/37_rostered_vs_worked_hours_by_work_center_shift.sql" in workforce_coverage_case_text
    assert "managerial/38_absence_rate_by_work_location_job_family_month.sql" in workforce_coverage_case_text
    assert "managerial/39_overtime_approval_coverage_and_concentration.sql" in workforce_coverage_case_text
    assert "managerial/41_late_arrival_early_departure_by_shift_department.sql" in workforce_coverage_case_text
    assert "## Recommended Query Sequence" not in workforce_coverage_case_text

    working_capital_case_text = Path("docs/analytics/cases/working-capital-and-cash-conversion-case.md").read_text(encoding="utf-8")
    assert "## The Problem to Solve" in working_capital_case_text
    assert "## What You Need to Develop" in working_capital_case_text
    assert "## Step-by-Step Walkthrough" in working_capital_case_text
    assert "## Optional Excel Follow-Through" in working_capital_case_text
    assert "## Wrap-Up Questions" in working_capital_case_text
    assert "financial/19_working_capital_bridge_by_month.sql" in working_capital_case_text
    assert "financial/02_ar_aging_open_invoices.sql" in working_capital_case_text
    assert "financial/15_customer_deposits_and_unapplied_cash_aging.sql" in working_capital_case_text
    assert "financial/03_ap_aging_open_invoices.sql" in working_capital_case_text
    assert "financial/20_cash_conversion_timing_review.sql" in working_capital_case_text
    assert "financial/09_payroll_liability_rollforward.sql" in working_capital_case_text
    assert "financial/11_payroll_cash_payments_and_remittances.sql" in working_capital_case_text
    assert "financial/12_accrued_expense_rollforward.sql" in working_capital_case_text
    assert "## Recommended Query Sequence" not in working_capital_case_text

    statement_bridge_case_text = Path("docs/analytics/cases/financial-statement-bridge-case.md").read_text(encoding="utf-8")
    assert "## The Problem to Solve" in statement_bridge_case_text
    assert "## What You Need to Develop" in statement_bridge_case_text
    assert "## Step-by-Step Walkthrough" in statement_bridge_case_text
    assert "## Optional Excel Follow-Through" in statement_bridge_case_text
    assert "## Wrap-Up Questions" in statement_bridge_case_text
    assert "financial/04_trial_balance_by_period.sql" in statement_bridge_case_text
    assert "financial/05_journal_and_close_cycle_review.sql" in statement_bridge_case_text
    assert "financial/06_control_account_reconciliation.sql" in statement_bridge_case_text
    assert "financial/16_retained_earnings_and_close_entry_impact.sql" in statement_bridge_case_text
    assert "financial/39_annual_income_to_equity_bridge.sql" in statement_bridge_case_text
    assert "financial/40_post_close_profit_and_loss_leakage_review.sql" in statement_bridge_case_text
    assert "financial/42_annual_net_revenue_bridge.sql" in statement_bridge_case_text
    assert "financial/43_invoice_revenue_cutoff_exception_summary.sql" in statement_bridge_case_text
    assert "financial/44_invoice_revenue_cutoff_exception_trace.sql" in statement_bridge_case_text
    assert "## Recommended Query Sequence" not in statement_bridge_case_text

    pricing_case_text = Path("docs/analytics/cases/pricing-and-margin-governance-case.md").read_text(encoding="utf-8")
    assert "## The Problem to Solve" in pricing_case_text
    assert "## What You Need to Develop" in pricing_case_text
    assert "## Step-by-Step Walkthrough" in pricing_case_text
    assert "## Optional Excel Follow-Through" in pricing_case_text
    assert "## Wrap-Up Questions" in pricing_case_text
    assert "financial/25_price_realization_vs_list_by_segment_customer_region_collection_style.sql" in pricing_case_text
    assert "financial/26_gross_margin_impact_of_promotions_vs_nonpromotion_sales.sql" in pricing_case_text
    assert "managerial/48_collection_revenue_margin_before_after_promotions.sql" in pricing_case_text
    assert "managerial/49_customer_specific_pricing_concentration_and_dependency.sql" in pricing_case_text
    assert "managerial/47_sales_rep_override_rate_and_discount_dispersion.sql" in pricing_case_text
    assert "managerial/50_monthly_price_floor_pressure_and_override_concentration.sql" in pricing_case_text
    assert "audit/51_override_approval_completeness_review.sql" in pricing_case_text
    assert "## Recommended Query Sequence" not in pricing_case_text

    o2c_text = Path("docs/processes/o2c.md").read_text(encoding="utf-8")
    assert "## Analytical Subsections" in o2c_text
    assert "## Returns, Credits, and Refunds" in o2c_text
    assert not Path("docs/processes/o2c-returns-credits-refunds.md").exists()

    p2p_text = Path("docs/processes/p2p.md").read_text(encoding="utf-8")
    assert "## Analytical Subsections" in p2p_text
    assert "direct accrued-service settlement" in p2p_text
    assert "PurchaseInvoiceLine.GoodsReceiptLineID" in p2p_text
    assert "PurchaseInvoiceLine.AccrualJournalEntryID" in p2p_text
    assert "DisbursementPayment" in p2p_text

    payroll_text = Path("docs/processes/payroll.md").read_text(encoding="utf-8")
    assert "## Analytical Subsections" in payroll_text
    assert "### Time, Attendance, and Approved Hours" in payroll_text
    assert "TimeClockEntry" in payroll_text
    assert "TimeClockPunch" in payroll_text
    assert "AttendanceException" in payroll_text
    assert "PayrollRegister" in payroll_text
    assert "PayrollLiabilityRemittance" in payroll_text
    assert not Path("docs/processes/time-clocks.md").exists()

    manufacturing_text = Path("docs/processes/manufacturing.md").read_text(encoding="utf-8")
    assert "## Analytical Subsections" in manufacturing_text
    assert "### Planning, MRP, and Rough-Cut Capacity" in manufacturing_text
    assert "MaterialRequirementPlan" in manufacturing_text
    assert "RoughCutCapacityPlan" in manufacturing_text
    assert "WorkOrderOperationSchedule" in manufacturing_text
    assert "LaborTimeEntry" in manufacturing_text
    assert "WorkOrderClose" in manufacturing_text

    manual_close_text = Path("docs/processes/manual-journals-and-close.md").read_text(encoding="utf-8")
    assert "## Analytical Subsections" in manual_close_text
    assert "### Accrual Estimate to AP Settlement" in manual_close_text
    assert "## Opening Balance and Year-End Close" in manual_close_text
    assert "PurchaseInvoiceLine.AccrualJournalEntryID" in manual_close_text
    assert "Accrual Adjustment" in manual_close_text
    assert "Year-End Close" in manual_close_text
