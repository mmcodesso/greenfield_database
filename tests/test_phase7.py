from dataclasses import replace
from pathlib import Path
import sqlite3
from zipfile import ZipFile

from openpyxl import load_workbook

from generator_dataset.anomalies import inject_anomalies
from generator_dataset.exporters import export_csv_zip, export_excel, export_sqlite, export_support_excel
from generator_dataset.main import build_phase6
from generator_dataset.schema import SQLITE_INDEXES, TABLE_COLUMNS, TABLE_PRIMARY_KEYS
from generator_dataset.validations import validate_phase7


def _sqlite_indexes(connection: sqlite3.Connection, table_name: str) -> dict[str, dict[str, object]]:
    indexes: dict[str, dict[str, object]] = {}
    for _, name, unique, _, _ in connection.execute(f"PRAGMA index_list('{table_name}')"):
        columns = [
            row[2]
            for row in connection.execute(f"PRAGMA index_info('{name}')")
        ]
        indexes[str(name)] = {
            "columns": columns,
            "unique": bool(unique),
        }
    return indexes


def test_phase7_anomalies_and_exports(tmp_path: Path) -> None:
    context = build_phase6()
    context.settings = replace(
        context.settings,
        export_support_excel=True,
        export_csv_zip=True,
        sqlite_path=str(tmp_path / "CharlesRiver.sqlite"),
        excel_path=str(tmp_path / "CharlesRiver.xlsx"),
        support_excel_path=str(tmp_path / "CharlesRiver_support.xlsx"),
        csv_zip_path=str(tmp_path / "CharlesRiver_csv.zip"),
    )

    inject_anomalies(context)
    results = validate_phase7(context)
    export_sqlite(context)
    export_excel(context)
    export_support_excel(context)
    export_csv_zip(context)

    assert results["exceptions"] == []
    assert results["anomaly_count"] > 0
    assert results["gl_balance"]["exception_count"] == 0
    assert (tmp_path / "CharlesRiver.sqlite").exists()
    assert (tmp_path / "CharlesRiver.xlsx").exists()
    assert (tmp_path / "CharlesRiver_support.xlsx").exists()
    assert (tmp_path / "CharlesRiver_csv.zip").exists()

    with sqlite3.connect(tmp_path / "CharlesRiver.sqlite") as connection:
        exported_tables = {
            row[0]
            for row in connection.execute("SELECT name FROM sqlite_master WHERE type = 'table'")
        }
        dataset_tables = {
            row[0]
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table' AND name NOT LIKE 'sqlite_%'"
            )
        }
    assert "AnomalyLog" not in exported_tables
    assert "ValidationSummary" not in exported_tables
    assert dataset_tables == set(TABLE_COLUMNS)

    with sqlite3.connect(tmp_path / "CharlesRiver.sqlite") as connection:
        for table_name in sorted(dataset_tables):
            table_info = connection.execute(f"PRAGMA table_info('{table_name}')").fetchall()
            pk_columns = [row[1] for row in table_info if row[5]]
            assert pk_columns == [TABLE_PRIMARY_KEYS[table_name]]

        for table_name, index_definitions in SQLITE_INDEXES.items():
            indexes = _sqlite_indexes(connection, table_name)
            for index_definition in index_definitions:
                assert index_definition.name in indexes
                assert indexes[index_definition.name]["columns"] == list(index_definition.columns)
                assert indexes[index_definition.name]["unique"] is index_definition.unique

        purchase_invoice_indexes = _sqlite_indexes(connection, "PurchaseInvoice")
        assert not any(
            index_data["unique"] and index_data["columns"] == ["InvoiceNumber"]
            for index_data in purchase_invoice_indexes.values()
        )

        disbursement_indexes = _sqlite_indexes(connection, "DisbursementPayment")
        assert not any(
            index_data["unique"] and index_data["columns"] == ["CheckNumber"]
            for index_data in disbursement_indexes.values()
        )

    dataset_workbook = load_workbook(tmp_path / "CharlesRiver.xlsx")
    support_workbook = load_workbook(tmp_path / "CharlesRiver_support.xlsx")

    assert "AnomalyLog" not in dataset_workbook.sheetnames
    assert "ValidationStages" not in dataset_workbook.sheetnames
    for sheet_name in ["Account", "SalesOrder", "GLEntry"]:
        assert len(dataset_workbook[sheet_name].tables) >= 1

    assert {"Overview", "AnomalyLog", "ValidationStages", "ValidationChecks", "ValidationExceptions"}.issubset(
        set(support_workbook.sheetnames)
    )
    for sheet_name in ["Overview", "AnomalyLog", "ValidationStages", "ValidationChecks", "ValidationExceptions"]:
        assert len(support_workbook[sheet_name].tables) >= 1

    with ZipFile(tmp_path / "CharlesRiver_csv.zip") as archive:
        zip_members = set(archive.namelist())
    assert zip_members == {f"{table_name}.csv" for table_name in context.tables}
